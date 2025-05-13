import math
import openpyxl

class Student():
    def __init__(self, name, hw, test):
        self.name = name
        self.hw = hw
        self.tests = test
        self.hlTest = 0
        self.hTest = 0
        self.hwGrade = 0
        self.quarterAvg = 0
        self.gradedHw = []
        self.gradedTests = []
        self.letters = {}

    def update(self):
        self.gradeHw()
        self.hwAverage()
        self.gradeTests()
        self.hlTestGrade(w1, w2)
        self.qAverage(w3, w4)
        self.letterGrade()

    def gradeHw(self):
        self.gradedHw = [self.hw[i] / self.maxHwGrades[i] for i in range(len(self.hw))]
        
    def product(self, a1, a2):
        return [a1[i] * a2[i] for i in range(len(a1))]
    
    def hwAverage(self):
        if 0 in self.hw:
            hwPoints = [self.hw[i] * self.hwWeights[i] for i in range(len(self.hw))]
            maxHwPoints = [self.maxHwGrades[i] * self.hwWeights[i] for i in range(len(self.hw))]
            self.hwGrade = sum(hwPoints) / sum(maxHwPoints)
            return self.hwGrade
        else:
            self.hwGrade = 1
            return 1

    def gradeTests(self):
        self.gradedTests = [self.tests[i] / self.maxTestGrades[i] for i in range(len(self.tests))]
        
    def testAverage(self):
        testScores = [self.tests[i] * self.testWeights[i] for i in range(len(self.tests))]
        maxTestScores = [self.maxTestGrades[i] * self.testWeights[i] for i in range(len(self.maxTestGrades))]
        return sum(testScores) / sum(maxTestScores)

    def hlTestGrade(self, w1, w2):
        tw = self.testWeights
        tw[-1] = w1
        g1 = self.testAverage()
        tw[-1] = w2
        g2 = self.testAverage()
        self.hlTest = g1, g2
        self.hTest = max(g1, g2)
        return g1, g2

    def qAverage(self, w1, w2):
        tw = w1 + w2
        w1 /= tw
        w2 /= tw
        self.quarterAvg = self.hwGrade * w1 + self.hTest *  w2
        return self.quarterAvg

    def letterGrade(self):
        hTest = letter(self.hTest)
        hwGrade = letter(self.hwGrade)
        quarterAvg = letter(self.quarterAvg)
        gradedHw = list(map(letter, self.gradedHw))
        gradedTests = list(map(letter, self.gradedTests))
        self.letters = {"hTest": hTest,
                        "hwGrade": hwGrade,
                        "quarterAvg": quarterAvg,
                        "gradedHw": gradedHw,
                        "gradedTests": gradedTests}
        return self.letters
        

def arithMean(arr):
    return sum(arr) / len(arr)

def geoMean(arr):
    prod = 1
    for i in arr:
        prod *= i
    return prod ** (1 / len(arr))

def harmMean(arr):
    if 0 in arr:
        return 0
    return len(arr) / sum(map(lambda n: 1 / n, arr))

def median(arr):
    arr.sort()
    n = len(arr) / 2
    a, b = math.floor(n), math.ceil(n)
    return (arr[a] + arr[b]) / 2

def mode(arr):
    s = {}
    for i in arr:
        if i in s:
            s[i] += 1
        else:
            s[i] = 1
    m = 0
    n = -1
    for i in s.items():
        if i[1] > m:
            m = i[1]
            n = i[0]
    return n

def standardDev(arr):
    m = arithMean(arr)
    return (sum([(i - m) ** 2 for i in arr]) / len(arr)) ** 0.5

def countLetters(arr):
    l = [0, 0, 0, 0, 0]
    for g in arr:
        if g == "A":
            l[0] += 1
        elif g == "B":
            l[1] += 1
        elif g == "C":
            l[2] += 1
        elif g == "D":
            l[3] += 1
        elif g == "F":
            l[4] += 1
    return l

def dist(students):
    #getting data
    hw = [student.gradedHw for student in students]
    tests = [student.gradedTests for student in students]
    qAvgs = [student.quarterAvg for student in students]
    avgHw = [student.hwGrade for student in students]
    avgTests = [student.hTest for student in students]

    #same assignments are grouped together
    hw = [[hw[i][j] for i in range(len(hw))] for j in range(len(hw[0]))]
    tests = [[tests[i][j] for i in range(len(tests))] for j in range(len(tests[0]))]
    #arr -> list of stats
    f = lambda n: (arithMean(n), geoMean(n), harmMean(n), median(n), mode(n), standardDev(n))
    hw = list(map(f, hw))
    tests = list(map(f, tests))
    qAvgs = f(qAvgs)
    avgHw = f(avgHw)
    avgTests = f(avgTests)

    hTestLetters = []
    avgHwLetters = []
    qAvgLetters = []
    tHwLetters = [[] for i in range(len(students[0].hw))]
    tTestLetters = [[] for i in range(len(students[0].tests))]
    #getting grades for letter counting
    for student in students:
        l = student.letters
        hTest = l["hTest"]
        hTestLetters.append(hTest)
        hwGrade = l["hwGrade"]
        avgHwLetters.append(hwGrade)
        quarterAvg = l["quarterAvg"]
        qAvgLetters.append(quarterAvg)
        gradedHw = l["gradedHw"]
        for i in range(len(gradedHw)):
            tHwLetters[i].append(gradedHw[i])
        gradedTests = l["gradedTests"]
        for i in range(len(gradedTests)):
            tTestLetters[i].append(gradedTests[i])

    hTestLetters = countLetters(hTestLetters)
    avgHwLetters = countLetters(avgHwLetters)
    qAvgLetters = countLetters(qAvgLetters)
    tHwLetters = list(map(countLetters, tHwLetters))
    tTestLetters = list(map(countLetters, tTestLetters))
    letterGrades = {"hTestLetters": hTestLetters,
                    "avgHwLetters": avgHwLetters,
                    "qAvgLetters": qAvgLetters,
                    "tHwLetters": tHwLetters,
                    "tTestLetters": tTestLetters}
        
    return {"hw": hw,
            "tests": tests,
            "qAvgs": qAvgs,
            "avgHw": avgHw,
            "avgTests": avgTests,
            "letterGrades": letterGrades}

def letter(n):
    n = round(n, 2)
    if n >= cutoffs[0]:
        return "A"
    if n >= cutoffs[1]:
        return "B"
    if n >= cutoffs[2]:
        return "C"
    if n >= cutoffs[3]:
        return "D"
    return "F"

if __name__ == "__main__":
    #input
    students = []
    #grade letter cutoffs
    with open('data2.csv', 'r') as file:
        data = file.read().split('\n')
        data[0] = data[0].split(",,,")
        data[1] = data[1].split(",,,")
        maxHwGrades = list(map(float, data[0][0].split(",")))
        maxTestGrades = list(map(float, data[0][1].split(",")))
        hwWeights = list(map(float, data[1][0].split(",")))
        testWeights = list(map(float, data[1][1].split(",")))[:-2] + [0]
        #grade letter cutoffs
        cutoffs = list(map(float, data[2].split(",")[:-1]))
        testCutoff = float(data[2].split(",")[-1])
        dMode = int(data[4])
        #final weights
        w1, w2 = map(float, data[1][1].split(",")[-2:])
        #test & hw weights
        w3, w4 = map(float, data[3].split(","))
        for i in data[6:]:
            i = i.split(',,,')
            i = list(map(lambda n: n.split(','), i))
            if not i[0][0]:
                continue
            students.append(Student(i[0][0], list(map(float, i[0][1:])), list(map(float, i[1]))))
            students[-1].maxHwGrades = maxHwGrades
            students[-1].hwWeights = hwWeights
            students[-1].maxTestGrades = maxTestGrades
            students[-1].testWeights = testWeights

    #updates values for students
    for student in students:
        student.update()
    d = dist(students)
    file = openpyxl.load_workbook("./gradebook.xlsx")
    file.create_sheet(str(int(file.sheetnames[-1]) + 1))
    gb = file[file.sheetnames[-1]]

    warningColor = openpyxl.styles.PatternFill(start_color='DD1111', end_color='DD1111', fill_type='solid')
    row = 1
    col = 1
    #fill in hw values
    gb.cell(row, col).value = "Name"
    col += 1
    for i in range(len(students[0].hw)):
        gb.cell(row, col).value = f"Homework #{i + 1}"
        col += 1
    gb.cell(row, col).value = "Missed HW Assignments"
    col += 1
    gb.cell(row, col).value = "HW Average"
    col += 1
    #fill in test data
    for i in range(len(students[0].tests)):
        gb.cell(row, col).value = f"Test #{i + 1}"
        col += 1
    gb.cell(row, col).value = "Failed Tests"
    col += 1
    gb.cell(row, col).value = f"Test Average with \nFinal Weight {w1}"
    col += 1
    gb.cell(row, col).value = f"Test Average with \nFinal Weight {w2}"
    col += 1
    gb.cell(row, col).value = "Test Averages"
    col += 1
    #final quarter grade
    gb.cell(row, col).value = "Quarter Averages"
    row += 1
    for i in range(len(students)):
        col = 1
        gb.cell(row, col).value = student.name
        col += 1
        student = students[i]
        #fill in hw grades
        for j in range(len(student.hw)):
            if dMode == 0:
                gb.cell(row, col).value = f"{student.hw[j]} {round(student.gradedHw[j] * 100)}% {student.letters['gradedHw'][j]}"
            elif dMode == 1:
                gb.cell(row, col).value = student.hw[j]
            elif dMode == 2:
                gb.cell(row, col).value = f"{round(student.gradedHw[j] * 100)}%"
            else:
                gb.cell(row, col).value = student.letters['gradedHw'][j]
            if student.hw[j] == 0:
                gb.cell(row, col).fill = warningColor
            col += 1
        gb.cell(row, col).value = student.hw.count(0)
        col += 1
        if dMode == 0:
            gb.cell(row, col).value = f"{round(student.hwGrade * 100)}% {student.letters['hwGrade']}"
        elif dMode == 3:
            gb.cell(row, col).value = student.letters['hwGrade']
        else:
            gb.cell(row, col).value = f"{round(student.hwGrade * 100)}%"
        col += 1
        #filling in test grades and failed tests
        for j in range(len(student.tests)):
            if dMode == 0:
                gb.cell(row, col).value = f"{student.tests[j]} {round(student.gradedTests[j] * 100)}% {student.letters['gradedTests'][j]}"
            elif dMode == 1:
                gb.cell(row, col).value = student.tests[j]
            elif dMode == 2:
                gb.cell(row, col).value = f"{round(student.gradedTests[j] * 100)}%"
            else:
                gb.cell(row, col).value = student.letters['gradedTests'][j]
            if student.gradedTests[j] < testCutoff:
                gb.cell(row, col).fill = warningColor
            col += 1
        gb.cell(row, col).value = len(list(filter(lambda n: n < testCutoff, student.gradedTests)))
        col += 1
        gb.cell(row, col).value = f"{round(student.hlTest[0] * 100)}%"
        col += 1
        gb.cell(row, col).value = f"{round(student.hlTest[1] * 100)}%"
        col += 1
        if dMode == 0:
            gb.cell(row, col).value = f"{round(student.hTest * 100)}% {student.letters['hTest']}"
            col += 1
            gb.cell(row, col).value = f"{round(student.quarterAvg * 100)}% {student.letters['quarterAvg']}"
            row += 1
        elif dMode == 3:
            gb.cell(row, col).value = student.letters['hTest']
            col += 1
            gb.cell(row, col).value = student.letters['quarterAvg']
            row += 1
        else:
            gb.cell(row, col).value = f"{round(student.hTest * 100)}%"
            col += 1
            gb.cell(row, col).value = f"{round(student.quarterAvg * 100)}%"
            row += 1

    #stats: mean, median, mode, standard deviation
    row += 1
    names = ["Arithmatic Mean",
             "Geometric Mean",
             "Harmonic Mean",
             "Median",
             "Mode",
             "Standard Deviation"]
    for i in range(6):
        row += 1
        col = 1
        gb.cell(row, col).value = names[i]
        col += 1
        for j in range(len(d["hw"])):
            gb.cell(row, col).value = f'{round(d["hw"][j][i] * 100)}%'
            col += 1
        col += 1
        gb.cell(row, col).value = f"{round(d['avgHw'][i] * 100)}%"
        col += 1
        for j in range(len(d["tests"])):
            gb.cell(row, col).value = f'{round(d["tests"][j][i] * 100)}%'
            col += 1
        col += 1
        col += 1
        col += 1
        gb.cell(row, col).value = f"{round(d['avgTests'][i] * 100)}%"
        col += 1
        gb.cell(row, col).value = f"{round(d['qAvgs'][i] * 100)}%"

    row += 2
    #filling in letter grades
    lg = ["A", "B", "C", "D", "F"]
    for i in range(5):
        col = 1
        gb.cell(row, col).value = f"# of {lg[i]}'s:"
        col += 1
        l = d["letterGrades"]
        for j in range(len(l["tHwLetters"])):
            gb.cell(row, col).value = l["tHwLetters"][j][i]
            col += 1
        col += 1
        gb.cell(row, col).value = l["avgHwLetters"][i]
        col += 1
        for j in range(len(l["tTestLetters"])):
            gb.cell(row, col).value = l["tTestLetters"][j][i]
            col += 1
        col += 1
        col += 1
        col += 1
        gb.cell(row, col).value = l["hTestLetters"][i]
        col += 1
        gb.cell(row, col).value = l["qAvgLetters"][i]
        row += 1

    #style changes
    alignment = openpyxl.styles.Alignment(horizontal = "left", vertical = "center")

    for col in gb.columns:
        mSize = 0
        for cell in col:
            if cell.value != None and cell.row != 1:
                cell.alignment = alignment
                mSize = max(mSize, len(str(cell.value)))
        gb.column_dimensions[col[0].column_letter].width = max(mSize + 3, 10)
    alignment = openpyxl.styles.Alignment(textRotation = 90, horizontal = "center", vertical = "center")
    gb.row_dimensions[1].height = 150
    for row in gb.rows:
        for cell in row:
            cell.alignment = alignment
        break
    
    bgColor = openpyxl.styles.PatternFill(start_color='222222', end_color='222222', fill_type='solid')
    textColor = openpyxl.styles.Color("00EEEEEE")
    font = openpyxl.styles.fonts.Font(name = "Aptos Narrow", color = textColor)
    borderColor = openpyxl.styles.Color("00BBBBBB")
    side = openpyxl.styles.borders.Side(style = "thin", color = borderColor)
    border = openpyxl.styles.borders.Border(left = side, right = side, top = side, bottom = side)
    for i in range(1, len(students) + 116):
        for j in range(1, len(student.hw) + len(student.tests) + 109):
            if gb.cell(i, j).fill != warningColor:
                gb.cell(i, j).fill = bgColor
    for i in range(1, len(students) + 16):
        if gb.cell(i, 1).value != None:
            for j in range(1, len(student.hw) + len(student.tests) + 9):
                gb.cell(i, j).font = font
                gb.cell(i, j).border = border
                if gb.cell(i, j).value == None:
                    gb.cell(i, j).value = "N/A"
    #save file
    file.save("gradebook.xlsx")
